import os
from huggingface_hub import login
from transformers import AutoTokenizer
import transformers
import torch
import openpyxl

# Replace with your Hugging Face API token
login(token='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')

model = "meta-llama/Llama-2-7b-chat-hf"

tokenizer = AutoTokenizer.from_pretrained(model)
pipeline = transformers.pipeline(
    "text-generation",
    model=model,
    torch_dtype=torch.float16,
    device_map="auto",
)

# Open the existing Excel file
current_dir = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(current_dir, "llama-2-7b_prompt_response.xlsx")
workbook = openpyxl.load_workbook(excel_file_path)
worksheet = workbook.active

# Get the maximum row index in the worksheet
max_row = worksheet.max_row

# Iterate through the rows and read prompts and responses
for row_index, row in enumerate(worksheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=2, values_only=True), start=2):
    prompt, response = row  # Read both prompt and response from the row
    
    print('prompt', prompt)
    # Skip rows where either the prompt or response is empty
    if prompt is None:
        continue

    # Generate a response for the prompt
    sequences = pipeline(
        prompt,
        do_sample=True,
        top_k=10,
        num_return_sequences=1,
        eos_token_id=tokenizer.eos_token_id,
        max_length=6000,
    )
    print('sequences', sequences)
    # Update the response in the Excel file
    worksheet.cell(row=row_index, column=2, value=sequences[0]["generated_text"])

    # Increment the row index
    max_row += 1

# Save the updated Excel file
workbook.save(excel_file_path)

# Close the Excel workbook
workbook.close()
